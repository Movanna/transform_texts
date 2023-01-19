# This script is used for gathering statistics about
# all of the texts in this large edition.

# Each language version of a text has a pre-made file
# in the GitHub repository, containing a template.
# Editors can add content to any file.
# The script measures the content length of each file
# in the repo by counting the characters in the human-readable
# text (the content of the tags) using Beautiful Soup.

# The script combines the content length with data from
# the database and outputs ordered CSV:s containing info
# about each text.

# The data is organized into pivot tables by Pandas and the
# different sheets are styled with openpyxl. This way I get
# nice statistics over how many texts there are in each
# language, how long each text is, who the translator is,
# how many pages of text there are within each category
# (i.e. letters, articles), how many images there are for 
# each text or category, and which texts are still missing
# from the repo (content length 0). This would otherwise be
# hard to know, because there are thousands of texts and files
# and the texts mainly exist only as XML. Editors can now easily
# check which texts haven't yet been transcribed or translated
# and how long a certain text would be were it to be printed, without
# keeping lists themselves or updating a list each time they've
# finished a text.

from pathlib import Path
import psycopg2
import re
from bs4 import BeautifulSoup
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle

conn_db = psycopg2.connect(
    host="",
    database="",
    user="",
    port="",
    password=""
)
cursor = conn_db.cursor()

COLLECTIONS = [1, 2, 3, 4, 5, 6, 7, 8, 9]
SOURCE_FOLDER = "C:/../GitHub/leomechelin_files/documents/Delutgava_"
EXCEL_FOLDER = "documents/statistik/"

# create path object for folder from given filepath string
# save all paths to files found in this folder or subfolders in a list
def create_file_list(source_folder):
    path = Path(source_folder)
    file_list = []
    iterate_through_folders(path, file_list)
    return file_list

# iterate through folders recursively and append filepaths to list
def iterate_through_folders(path, file_list):
    for content in path.iterdir():
        if content.is_dir():
            iterate_through_folders(content, file_list)
        elif content.suffix == ".xml":
            file_list.append(content)

# publication_id and language are present in the file name
# folders are in the file path 
def extract_info_from_filename(file):
    file_name = file.stem
    main_folder = file.parts[8]
    if main_folder == "Brev" or main_folder == "Artiklar" or main_folder == "Lantdagen" or main_folder == "Verk" or main_folder == "Forelasningar":
        subfolder = file.parts[9]
    else:
        subfolder = None
    if main_folder == "Brev":
        correspondent_folder = file.parts[10]
    else:
        correspondent_folder = None
    # create file path string and shorten it
    file_path = file.as_posix().replace("C:/../GitHub/leomechelin_files/", "")
    file_data = (main_folder, subfolder, correspondent_folder, file_path)
    search_string = re.compile(r"(\w{2})_(\d+)$")
    match = re.search(search_string, file_name)
    # even if db value original_language contains several values,
    # the filename has just the first of them
    language = match.group(1)
    publication_id = match.group(2)
    return file_data, publication_id, language

# get the relevant info for the publication for this file
def fetch_db_data(publication_id, language):
    # these publications don't have separate manuscript files/data
    print(str(publication_id) + " " + language)
    if language == "sv" or language == "fi":
        fetch_query = """SELECT publication_group_id, text, original_publication_date, original_language FROM publication, translation_text WHERE publication.id = %s AND translation_text.language = %s AND publication.translation_id = translation_text.translation_id AND field_name = %s"""
        field_name = "name"
        values_to_insert = (publication_id, language, field_name)
    # these publications always have a manuscript file/data
    else:
        fetch_query = """SELECT publication_group_id, publication_manuscript.name, publication.original_publication_date, publication_manuscript.original_language FROM publication, publication_manuscript WHERE publication.id = %s AND publication_manuscript.original_language LIKE %s AND publication.id = publication_id"""
        # language value for these publications may contain several
        # languages, but files only have one registered language per file
        check_language = "%" + language + "%"
        values_to_insert = (publication_id, check_language)
    cursor.execute(fetch_query, values_to_insert)
    db_data = cursor.fetchone()
    (group_id, title, date, original_language) = db_data
    # if this file is the original language version of the files
    # for this publication, it isn't a translation
    # we need to separate these files from the ones that are meant
    # to be translated but just haven't been translated yet
    # (their translator value is None)
    # later on, we'll replace this temporary value with
    # a specific cell styling in our Excel report
    # since there may be more than one translator for a text
    # variable translator is a tuple containing a list of tuples
    if original_language == language or language in original_language:
        translator = ([("X", "X")],)
        db_data = db_data + translator
    else:
        # fetch the translator(s) for this text, if it has any
        fetch_query = """SELECT last_name, first_name FROM contributor, contribution WHERE contributor.id = contribution.contributor_id AND contribution.publication_id = %s AND text_language = %s AND contribution.deleted = %s"""
        deleted = 0
        values_to_insert = (publication_id, language, deleted)
        cursor.execute(fetch_query, values_to_insert)
        translators = cursor.fetchall()
        if translators != []:
            translator = (translators,)
            db_data = db_data + translator
        else:
            db_data = db_data + ([(None, None)],)
    # fetch the number of images for this publication
    # (in the db this is registered as "number_of_pages")
    # one publication may have several facsimiles, i.e. separate image units
    # we'll count them all
    # there are two or three language versions/files for each publication,
    # but only the original language version/file has images
    # the other texts/files are (at least in part) translations
    # with no own images
    # since all these versions share the same publication_id
    # register the number of images only once per publication
    # otherwise the total number of images will be wrong:
    # if original_language for publication_id 1 
    # is "de, sv", only count the images once, for the de-file,
    # not for the sv-file or fi-file
    images = None
    first_original_language = False
    if language in original_language:
        languages = original_language.split(", ")
        if language == languages[0]:
            first_original_language = True
    # if this is the original file, check possible images
    # else nr of images is 0
    images = 0
    if original_language == language or first_original_language is True:
        fetch_query = """SELECT number_of_pages FROM publication_facsimile_collection, publication_facsimile WHERE publication_facsimile.publication_id = %s AND publication_facsimile.publication_facsimile_collection_id = publication_facsimile_collection.id AND publication_facsimile_collection.deleted = %s"""
        deleted = 0
        values_to_insert = (publication_id, deleted)
        cursor.execute(fetch_query, values_to_insert)
        image_values = cursor.fetchall()
        # if there is one or more connected facsimile(s)
        if image_values != []:
            # there may be several facsimiles, count them all
            # facsimile_type 0 represents a link to another site
            # containing the images
            # in that case there are no image values to count, but we still need
            # a value in order to be able to calculate the sum of images
            # let's use a number that can get rounded to 0 and therefore won't
            # affect totals, but still differs from 0 and the other
            # image int values
            # later on, we'll replace this float value with
            # a specific cell styling in our Excel report
            for image_value in image_values:
                number_of_images = image_value[0]
                facsimile_type = image_value[1]
                if facsimile_type != 0 and number_of_images is not None:
                    images += image_value[0]
                if facsimile_type == 0:
                    images += 0.001
    db_data = db_data + (images,)
    return db_data

# read an xml file and return its content as a soup object
def read_xml(file):
    with file.open("r", encoding="utf-8-sig") as source_file:
        file_content = source_file.read()
        xml_soup = BeautifulSoup(file_content, "xml")
    return xml_soup

# check whether there is any content in the file, apart from
# the template
# strip newlines (otherwise they'll get counted as characters)
# and use space to join the text contents of all the elements
# this probably is the most accurate way to measure the length
# without further processing the string
# the content length isn't going to be 100% right, but good enough
# also calculate the estimated page value if the text were to
# be published as a printed book
def check_content(xml_soup):
    main_div = xml_soup.body.div
    if len(main_div.get_text(strip = True)) == 0:
        content_length = 0
        pages = 0
    else:
        content_length = len(main_div.get_text(" ", strip = True))
        # we will assume there are 2 500 characters including spaces
        # on a printed page
        pages = round(content_length / 2500, 1)
    return content_length, pages

# construct the url for each publication
def construct_url(publication_id, COLLECTION_ID):
    url = "https://digital_publishing_project/publication/" + str(COLLECTION_ID) + "/text/" + str(publication_id) + "/nochapter/not/infinite/nosong/searchtitle/established_sv&established_fi&facsimiles&manuscripts"
    return url

# make a list out of the values for each file
# add that list to the main list holding all files' values
def construct_list(file_data, publication_id, language, db_data, content_length, pages, url, stats_list):
    publication_info = []
    (main_folder, subfolder, correspondent_folder, file_path) = file_data
    (group_id, title, date, original_language, translators, images) = db_data
    # add the translator's name, or if there is no translator
    # leave this slot empty
    # variable translators is a list of tuples, so we have to get
    # each tuple in the list, and then the values in that tuple
    i = 0
    for tuple in translators:
        translator_last_name = tuple[0]
        if translator_last_name is None:
            translator = ""
            break
        translator_first_name = tuple[1]
        if i == 0:
            translator = translator_last_name + ", " + translator_first_name
        else:
            translator = translator + "; " + translator_last_name + ", " + translator_first_name
        i += 1
    # if the language of the file content is the document's
    # original language: make this explicit in the report
    if language == original_language or language in original_language:
        language = language + " (orig.)"
    if group_id is None:
        group_id = 100
    publication_info.append(publication_id)
    publication_info.append(group_id)
    publication_info.append(language)
    publication_info.append(date)
    publication_info.append(title)
    publication_info.append(content_length)
    publication_info.append(pages)
    publication_info.append(main_folder)
    publication_info.append(subfolder)
    publication_info.append(correspondent_folder)
    publication_info.append(translator)
    publication_info.append(images)
    publication_info.append(url)
    publication_info.append(file_path)
    stats_list.append(publication_info)
    return stats_list

def sort_stats_list(stats_list):
    # sort the list by these criteria:
    # main folder, subfolder, correspondent folder, date, group, id, language
    stats_list_1 = stats_list.copy()
    stats_list_1.sort(key = lambda row: (row[7], row[8], row[9], row[3], row[1], row[0], row[2]))
    return stats_list_1

# style the Excel sheets using openpyxl
# and replace some values with a certain style
def style_spreadsheet(spreadsheet_file_path):
    # open the newly created workbook
    workbook = openpyxl.load_workbook(filename = spreadsheet_file_path)
    # header style
    header = NamedStyle(name = "header")
    header.font = Font(bold = True, color = "081354")
    header.border = Border(bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center")
    # table style
    table_font = Font(color = "081354")
    table_border = Border(bottom = Side(border_style = "thin"))
    # loop through the sheets
    for name in workbook.sheetnames:
        sheet = workbook[name]
        # first row is the header, style it differently
        header_row = sheet[1]
        for cell in header_row:
            cell.style = header
        # do this only for sheet 1 (shortest sheet title)
        if len(sheet.title) < 10:
            # loop through translator values and replace value
            # "X, X" (which means that this is the original language file)
            # with a pattern fill, so you can easily spot those rows
            # knowing they don't lack the translator since they're originals
            r = 2
            for row in sheet.iter_rows():
                cell_to_fix = sheet.cell(row = r, column = 11)
                if cell_to_fix.value == "X, X":
                    cell_to_fix.value = ""
                    cell_to_fix.fill = PatternFill(fill_type = "lightTrellis")
                r += 1
            # loop through values for number of images and replace value 0
            # with a pattern fill if this is a row for something else than 
            # an original language file, because only original language files
            # can have images
            # also replace float values > 0 and < 1 with a string
            # because these values represent an external link and are
            # meant to be replaced and not to be displayed as such
            # (see function fetch_db_data)
            # finally change remaining floats to int in order to get rid of
            # possible decimal points > 0.000 indicating that a publication has
            # both images and a link (values like 4.001 should be displayed as 4)
            r = 2
            for row in sheet.iter_rows():
                orig = False
                cell_to_fix = sheet.cell(row = r, column = 12)
                cell_to_check = sheet.cell(row = r, column = 3)
                if cell_to_fix.value is not None:
                    if "(orig.)" in str(cell_to_check.value):
                        orig = True
                    if float(cell_to_fix.value) == 0 and orig is False:
                        cell_to_fix.value = ""
                        cell_to_fix.fill = PatternFill(fill_type = "lightTrellis")
                    elif 0 < float(cell_to_fix.value) < 1 and orig is True:
                        cell_to_fix.value = "extern länk"
                    elif float(cell_to_fix.value) > 1:
                        cell_to_fix.value = int(cell_to_fix.value)
                    else:
                        r += 1
                        continue
                r += 1
        # set column width and styles for the different types of sheets
        if len(sheet.title) < 10:
            sheet.column_dimensions["A"].width = 10
            sheet.column_dimensions["B"].width = 10
            sheet.column_dimensions["C"].width = 10
            sheet.column_dimensions["D"].width = 11
            sheet.column_dimensions["E"].width = 50
            sheet.column_dimensions["F"].width = 13
            sheet.column_dimensions["G"].width = 13
            sheet.column_dimensions["H"].width = 13
            sheet.column_dimensions["I"].width = 15
            sheet.column_dimensions["J"].width = 23
            sheet.column_dimensions["K"].width = 30
            sheet.column_dimensions["L"].width = 10
            sheet.column_dimensions["M"].width = 35
            sheet.column_dimensions["N"].width = 20
            # add gradient colours depending on value in column
            # "printed pages", ranging from orange for files
            # with no content length, through yellow to green
            color_scale_rule = ColorScaleRule(start_type = "num", start_value = 0, start_color = "FF9933",  mid_type = "num", mid_value = 5, mid_color = "FFF033", end_type = "num", end_value = 300, end_color = "97FF33")
            sheet.conditional_formatting.add("G2:G510000", color_scale_rule)
            # add similar gradient colours depending on value in
            # column "number of images"
            color_scale_rule_2 = ColorScaleRule(start_type = "num", start_value = 0, start_color = "FF9933",  mid_type = "num", mid_value = 3, mid_color = "FFF033", end_type = "num", end_value = 400, end_color = "97FF33")
            sheet.conditional_formatting.add("L2:L510000", color_scale_rule_2)
        elif len(sheet.title) > 30:
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 20
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 20
        else:
            sheet.column_dimensions["A"].width = 20
            sheet.column_dimensions["B"].width = 20
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 20
            # style cells in column A as long as they're not blank 
            diff_style = DifferentialStyle(border = table_border, font = table_font)
            rule = Rule(type="expression", dxf = diff_style)
            rule.formula = ["NOT(ISBLANK(A2:A100))"]
            sheet.conditional_formatting.add("A2:A100", rule)
        # make header row of each sheet stick when scrolling
        sheet.freeze_panes = "N2"
    workbook.save(spreadsheet_file_path)
    print("Workbook updated with styles.")

# for each file: extract data from the filename
# and fetch additional data from the db
# measure content length of the file
# and construct additional data
# make a list out of the values and append to main list
# then sort the main list
# use Pandas to create spreadsheet data and pivot tables
def create_data_and_tables(file_list, collection_id):
    stats_list = []
    for file in file_list:
        file_data, publication_id, language = extract_info_from_filename(file)
        db_data = fetch_db_data(publication_id, language)
        xml_soup = read_xml(file)
        content_length, pages = check_content(xml_soup)
        url = construct_url(publication_id, collection_id)
        stats_list = construct_list(file_data, publication_id, language, db_data, content_length, pages, url, stats_list)
    stats_list_sorted = sort_stats_list(stats_list)
    # use Pandas to create spreadsheet data and pivot tables
    # the data frame will be sheet 1
    df = pd.DataFrame(stats_list_sorted, columns = ["id", "grupp", "språk", "datum", "titel", "teckenmängd", "tryckta_sidor", "genre", "undermapp", "korrespondent", "översättare", "bildantal", "länk", "fil"])
    # earlier on we recorded some specific image data as floats
    # for the purpose of easily replacing it in sheet 1
    # (see function fetch_db_data)
    # for this df we want to get rid of those decimals as we want
    # to sum integers, or otherwise the result will be wrong
    df_images_as_int = pd.DataFrame(stats_list_sorted, columns = ["id", "grupp", "språk", "datum", "titel", "teckenmängd", "tryckta_sidor", "genre", "undermapp", "korrespondent", "översättare", "bildantal", "länk", "fil"])
    df_images_as_int = df_images_as_int.astype({"bildantal": int})
    # table_1 has both subtotals and totals, which we have to get
    # by concatenating two slightly different pivot tables
    # this is the number of pages per language per genre
    # and also the number of images per language per genre
    piv_1 = df_images_as_int.pivot_table(index = ["genre", "språk"], values = ["tryckta_sidor", "bildantal"], aggfunc = "sum", margins = True, margins_name = "summa")
    # this is the subtotal of the number of pages per genre
    # and also the subtotal of the number of images per language per genre
    piv_2 = piv_1.query("genre != 'summa'").groupby(level = 0).sum().assign(språk = "totalt").set_index("språk", append = True)
    table_1 = pd.concat([piv_1, piv_2]).sort_index()
    # table_2 is a pivot table of the total number of pages
    # for each language and also of how much
    # there is to translate into Finnish and Swedish
    piv_3 = df.pivot_table(index = "språk", values = "tryckta_sidor", aggfunc = "sum")
    # add new columns containing values calculated
    # from the content of newly created pivot table piv_3
    # i.e. add "pages requiring translation"
    piv_4 = piv_3.assign(att_översätta_till_fi = piv_3.query("språk != 'summa' and språk != 'fi' and språk != 'fi (orig.)' and språk != 'sv'"), att_översätta_till_sv = piv_3.query("språk != 'summa' and språk != 'sv' and språk != 'sv (orig.)' and språk != 'fi'")).fillna(0)
    # also find out the value of already translated texts
    # this query returns a whole row
    translated_into_fi = piv_4.query("språk == 'fi'")
    translated_into_sv = piv_4.query("språk == 'sv'")
    # the value we want is at position row 0, column 0
    # we may also get an empty DataFrame back, if nothing
    # has been translated yet
    if translated_into_fi.size != 0:
        translated_into_fi = translated_into_fi.iloc[0, 0]
    else:
        translated_into_fi = 0
    if translated_into_sv.size != 0:
        translated_into_sv = translated_into_sv.iloc[0, 0]
    else:
        translated_into_sv = 0
    piv_4 = piv_4.pivot_table(index = "språk", aggfunc = "sum", margins = True, margins_name = "summa")
    # change column order
    table_2 = piv_4.loc[:, ["tryckta_sidor", "att_översätta_till_fi", "att_översätta_till_sv"]]
    # append last row containing the values of already translated texts
    table_2.loc["redan översatt"] = ["", translated_into_fi, translated_into_sv]
    # table_3 is only for collections that contain letters
    # it's a pivot table of the total number of pages for the
    # different language versions of a correspondent's letters
    df_filtered = df.query("genre == 'Brev'")
    try:
        table_3 = pd.pivot_table(df_filtered, values = "tryckta_sidor", index = ["korrespondent", "språk"], columns = "undermapp", aggfunc = "sum", margins = True, margins_name = "summa")
    except:
        table_3 = None
    # construct the sheet names
    # there are 3 or 4 sheets for each collection:
    # the data frame and the pivot tables
    df_sheet_name = "utg. " + str(collection_id)
    table_1_sheet_name = "utg. " + str(collection_id) + ", sidor per genre"
    table_2_sheet_name = "utg. " + str(collection_id) + ", sidor per språk"
    table_3_sheet_name = "utg. " + str(collection_id) + ", sidor per korrespondent"
    # use current month for constructing the workbook title
    # since this is a monthly report
    file_date = datetime.datetime.now()
    file_date = file_date.strftime("%m") + "_" + file_date.strftime("%Y")
    spreadsheet_file_path = EXCEL_FOLDER + "Rapport_" + file_date + ".xlsx"
    # check if the Excel workbook has already been created
    # if it has, append to the existing one
    path = Path(spreadsheet_file_path)
    if path.is_file():
        with pd.ExcelWriter(spreadsheet_file_path, engine = "openpyxl", mode = "a") as writer:
            df.to_excel(writer, sheet_name = df_sheet_name, index = False)
            table_1.to_excel(writer, sheet_name = table_1_sheet_name)
            table_2.to_excel(writer, sheet_name = table_2_sheet_name)
            if table_3 is not None:
                table_3.to_excel(writer, sheet_name = table_3_sheet_name)
    else:
        with pd.ExcelWriter(spreadsheet_file_path, engine = "openpyxl") as writer:
            df.to_excel(writer, sheet_name = df_sheet_name, index = False)
            table_1.to_excel(writer, sheet_name = table_1_sheet_name)
            table_2.to_excel(writer, sheet_name = table_2_sheet_name)
            if table_3 is not None:
                table_3.to_excel(writer, sheet_name = table_3_sheet_name)
    return spreadsheet_file_path

def main():
    # create a list of all files in a collection
    # then create spreadsheet data for the texts in
    # each collection
    # finally style the spreadsheet
    for collection_id in COLLECTIONS:
        source_folder = SOURCE_FOLDER + str(collection_id)
        file_list = create_file_list(source_folder)
        spreadsheet_file_path = create_data_and_tables(file_list, collection_id)
    print("Workbook " + str(spreadsheet_file_path) + " created.")
    style_spreadsheet(spreadsheet_file_path)

main()